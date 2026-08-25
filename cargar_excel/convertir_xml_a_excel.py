#!/usr/bin/env python3
"""
xml2excel.py — Convierte un archivo XML a Excel (.xlsx) sin perder campos.

Diseño:
  - Detecta automaticamente el elemento "registro" (el que se repite) o se
    puede forzar con --record-path.
  - Modo 'wide'      : una fila por registro; jerarquia aplanada a columnas
                       tipo  cliente/direccion/@codigo  o  items/item[2]/precio
  - Modo 'relational': normaliza las listas anidadas en hojas separadas
                       unidas por _id / _parent_id (sin columnas repetidas).
  - Modo 'both'      : ambas cosas en el mismo libro.

Garantias de "no perder datos":
  - Se exportan atributos (@nombre), texto de nodos hoja, texto mixto (#text),
    nodos vacios (columna presente con valor vacio) y namespaces (opcional).
  - Se respetan los limites de Excel: 1.048.576 filas por hoja (se parte en
    varias), 32.767 caracteres por celda (se trunca y se avisa en la hoja _info),
    caracteres de control ilegales (se eliminan), nombres de hoja de 31 chars.
  - Por defecto TODO se escribe como texto: los IDs con ceros a la izquierda,
    los numeros largos y las fechas ambiguas NO se corrompen. Usa --infer-types
    si prefieres tipos nativos (asumiendo el riesgo).
  - Los valores que empiezan por = + - @ se escriben como texto plano para
    evitar inyeccion de formulas al abrir el archivo.

Uso:
    python xml2excel.py entrada.xml -o salida.xlsx
    python xml2excel.py entrada.xml --mode relational --keep-namespaces
    python xml2excel.py entrada.xml --record-path "root/catalogo/producto"
    python xml2excel.py entrada.xml --list-paths      # inspeccionar antes
    python xml2excel.py grande.xml  --stream registro # archivos enormes
"""

from __future__ import annotations

import argparse
import datetime as dt
import re
import sys
import xml.etree.ElementTree as ET
from collections import Counter, OrderedDict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MAX_CELL_LEN = 32767
MAX_ROWS_PER_SHEET = 1_048_575  # 1.048.576 menos la fila de cabecera
WARNINGS: list[str] = []
INTERACTIVO = False


# --------------------------------------------------------------------------
# Nombres y namespaces
# --------------------------------------------------------------------------
def make_namer(keep_ns: bool, nsmap: dict[str, str]):
    """Devuelve una funcion tag -> nombre legible."""

    def name(tag: str) -> str:
        if not isinstance(tag, str) or not tag.startswith("{"):
            return tag
        uri, local = tag[1:].split("}", 1)
        if not keep_ns:
            return local
        prefix = nsmap.get(uri)
        return f"{prefix}:{local}" if prefix else f"{{{uri}}}{local}"

    return name


def parse_with_namespaces(path: str):
    """Parsea el XML y recupera el mapa uri -> prefijo (ET lo pierde si no)."""
    nsmap: dict[str, str] = {}
    root = None
    for event, payload in ET.iterparse(path, events=("start-ns", "start")):
        if event == "start-ns":
            prefix, uri = payload
            nsmap.setdefault(uri, prefix or "ns")
        else:
            if root is None:
                root = payload
                break
    tree = ET.parse(path)
    return tree.getroot(), nsmap


# --------------------------------------------------------------------------
# Recorrido / aplanado
# --------------------------------------------------------------------------
def node_text(elem) -> str:
    """Texto propio + colas de los hijos (contenido mixto), normalizado."""
    parts = [elem.text or ""]
    for child in elem:
        parts.append(child.tail or "")
    text = "".join(parts)
    return re.sub(r"\s+", " ", text).strip()


def build_schema(records, name) -> dict:
    """Pre-analisis: garantiza columnas ESTABLES entre registros.

    - multi[ruta_rel]  : la etiqueta se repite en algun registro -> siempre [n]
    - mixed[ruta_rel]  : la etiqueta tiene hijos en algun registro -> texto en /#text
    """
    multi: set[str] = set()
    mixed: set[str] = set()

    def walk(elem, rel):
        children = [c for c in elem if isinstance(c.tag, str)]
        if children and rel:
            mixed.add(rel)
        counts = Counter(name(c.tag) for c in children)
        for child in children:
            tag = name(child.tag)
            child_rel = f"{rel}/{tag}" if rel else tag
            if counts[tag] > 1:
                multi.add(child_rel)
            walk(child, child_rel)

    for rec in records:
        walk(rec, "")
    return {"multi": multi, "mixed": mixed}


def flatten(elem, name, prefix: str = "", out: "OrderedDict[str, str]" = None,
            skip_paths: set[str] = frozenset(), abs_path: str = "",
            schema: dict | None = None, rel: str = "") -> "OrderedDict[str, str]":
    """Aplana un elemento a pares ruta -> valor.

    skip_paths: rutas absolutas de sub-elementos que se exportan en otra hoja
                (modo relacional) y por tanto no deben aplanarse aqui.
    schema:     resultado de build_schema(); estabiliza los nombres de columna.
    """
    if out is None:
        out = OrderedDict()
    multi = schema["multi"] if schema else set()
    mixed = schema["mixed"] if schema else set()

    def put(key: str, value: str) -> None:
        if key in out:                     # colision improbable: no se pisa nada
            i = 2
            while f"{key}#{i}" in out:
                i += 1
            key = f"{key}#{i}"
        out[key] = value

    for attr, value in elem.attrib.items():
        put(f"{prefix}@{name(attr)}", value)

    children = [c for c in elem if isinstance(c.tag, str)]  # ignora comentarios/PI
    text = node_text(elem)
    leaf_key = prefix[:-1] if prefix.endswith("/") else prefix

    is_mixed = bool(children) or (rel in mixed)
    if text:
        put(f"{prefix}#text" if is_mixed else (leaf_key or "#text"), text)
    elif not children and not elem.attrib:
        put((f"{prefix}#text" if is_mixed else leaf_key) or "#text", "")

    counts = Counter(name(c.tag) for c in children)
    seen: Counter = Counter()
    for child in children:
        tag = name(child.tag)
        child_abs = f"{abs_path}/{tag}" if abs_path else tag
        child_rel = f"{rel}/{tag}" if rel else tag
        if child_abs in skip_paths:
            continue
        seen[tag] += 1
        idx = f"[{seen[tag]}]" if (counts[tag] > 1 or child_rel in multi) else ""
        flatten(child, name, f"{prefix}{tag}{idx}/", out, skip_paths, child_abs,
                schema, child_rel)

    return out


# --------------------------------------------------------------------------
# Deteccion del elemento "registro"
# --------------------------------------------------------------------------
def path_stats(root, name):
    """Cuenta apariciones y profundidad de cada ruta absoluta de elementos."""
    counts: Counter = Counter()
    depth: dict[str, int] = {}

    def walk(elem, path, d):
        for child in elem:
            if not isinstance(child.tag, str):
                continue
            p = f"{path}/{name(child.tag)}"
            counts[p] += 1
            depth.setdefault(p, d + 1)
            walk(child, p, d + 1)

    root_path = name(root.tag)
    counts[root_path] = 1
    depth[root_path] = 0
    walk(root, root_path, 0)
    return counts, depth


def detect_record_path(counts: Counter, depth: dict[str, int]) -> str | None:
    # El registro es la ruta REPETIDA MAS SUPERFICIAL (no la mas frecuente:
    # los hijos anidados suelen repetirse mas veces que el propio registro).
    candidates = [p for p, c in counts.items() if c > 1]
    if not candidates:
        return None
    return min(candidates, key=lambda p: (depth[p], -counts[p], p))


def select_by_path(root, name, target: str) -> list:
    """Devuelve los elementos cuya ruta absoluta coincide con `target`."""
    found: list = []

    def walk(elem, path):
        if path == target:
            found.append(elem)
            return                          # no anidamos registros dentro de si mismos
        for child in elem:
            if isinstance(child.tag, str):
                walk(child, f"{path}/{name(child.tag)}")

    walk(root, name(root.tag))
    return found


# --------------------------------------------------------------------------
# Modo relacional
# --------------------------------------------------------------------------
def repeating_children(records, name, base_path: str) -> dict[str, str]:
    """Rutas hijas que se repiten en algun registro -> se llevan a otra hoja."""
    repeated: dict[str, str] = {}

    def scan(elem, path):
        counts = Counter(name(c.tag) for c in elem if isinstance(c.tag, str))
        for tag, n in counts.items():
            if n > 1:
                repeated[f"{path}/{tag}"] = tag
        for child in elem:
            if isinstance(child.tag, str):
                scan(child, f"{path}/{name(child.tag)}")

    for rec in records:
        scan(rec, base_path)
    return repeated


def build_relational(records, name, base_path: str, schema: dict):
    """Devuelve {ruta_tabla: (columnas, filas)} normalizando las repeticiones."""
    repeated = set(repeating_children(records, name, base_path))
    tables: "OrderedDict[str, list[OrderedDict]]" = OrderedDict()
    tables[base_path] = []
    counters: Counter = Counter()

    def sub_tables(path: str) -> set[str]:
        """Tablas hijas inmediatas: repetidas bajo `path` sin otra tabla intermedia."""
        out = set()
        for p in repeated:
            if not p.startswith(path + "/"):
                continue
            if any(q != p and p.startswith(q + "/") and q.startswith(path + "/")
                   for q in repeated):
                continue
            out.add(p)
        return out

    def descend(elem, path, targets, row_id):
        """Localiza los elementos de las tablas hijas, aunque haya wrappers."""
        for child in elem:
            if not isinstance(child.tag, str):
                continue
            child_path = f"{path}/{name(child.tag)}"
            if child_path in targets:
                emit(child, child_path, row_id)
            else:
                descend(child, child_path, targets, row_id)

    def emit(elem, path, parent_id):
        counters[path] += 1
        row_id = f"{path.rsplit('/', 1)[-1]}-{counters[path]}"
        targets = sub_tables(path)
        rel_base = path[len(base_path) + 1:] if path != base_path else ""
        flat = flatten(elem, name, "", None, targets, path, schema, rel_base)
        row = OrderedDict()
        row["_id"] = row_id
        if parent_id is not None:
            row["_parent_id"] = parent_id
        row.update(flat)
        tables.setdefault(path, []).append(row)
        descend(elem, path, targets, row_id)

    for rec in records:
        emit(rec, base_path, None)

    result: "OrderedDict[str, tuple[list[str], list[OrderedDict]]]" = OrderedDict()
    for path, rows in tables.items():
        cols: list[str] = []
        for row in rows:
            for k in row:
                if k not in cols:
                    cols.append(k)
        result[path] = (cols, rows)
    return result


# --------------------------------------------------------------------------
# Escritura en Excel
# --------------------------------------------------------------------------
DANGEROUS = ("=", "+", "-", "@", "\t", "\r")
ISO_DT = re.compile(r"^\d{4}-\d{2}-\d{2}([T ]\d{2}:\d{2}(:\d{2}(\.\d+)?)?)?(Z|[+-]\d{2}:?\d{2})?$")
INT_RE = re.compile(r"^-?\d{1,15}$")
FLOAT_RE = re.compile(r"^-?\d+\.\d+([eE][+-]?\d+)?$")


def coerce(value: str):
    """Conversion conservadora a tipo nativo (solo con --infer-types)."""
    s = value.strip()
    if not s:
        return ""
    if INT_RE.match(s) and not re.match(r"^-?0\d", s):
        return int(s)
    if FLOAT_RE.match(s):
        return float(s)
    low = s.lower()
    if low in ("true", "false"):
        return low == "true"
    if ISO_DT.match(s):
        try:
            parsed = dt.datetime.fromisoformat(s.replace("Z", "+00:00"))
        except ValueError:
            return s
        if parsed.tzinfo is not None:
            # Excel no tiene tipo "datetime con zona horaria": convertirlo
            # implicaria perder el offset, asi que se conserva como texto.
            WARNINGS.append("Fechas con zona horaria conservadas como texto "
                            "(Excel no soporta offsets); ej: " + s)
            return s
        return parsed
    return s


def clean(value: str, path_hint: str) -> str:
    value = ILLEGAL_CHARACTERS_RE.sub("", value)
    if len(value) > MAX_CELL_LEN:
        WARNINGS.append(
            f"Valor truncado a {MAX_CELL_LEN} caracteres en '{path_hint}' "
            f"(longitud original: {len(value)})"
        )
        value = value[:MAX_CELL_LEN - 1] + "\u2026"
    return value


def write_sheet(wb: Workbook, title: str, columns: list[str], rows: list[dict],
                infer_types: bool) -> None:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="44546A")
    chunks = [rows[i:i + MAX_ROWS_PER_SHEET] for i in range(0, len(rows), MAX_ROWS_PER_SHEET)] or [[]]

    for n, chunk in enumerate(chunks, start=1):
        name = title if len(chunks) == 1 else f"{title[:27]}_{n}"
        ws = wb.create_sheet(unique_sheet_name(wb, name))
        ws.append(columns)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center", wrap_text=False)

        for row in chunk:
            ws.append([""] * len(columns))
            r = ws.max_row
            for c, col in enumerate(columns, start=1):
                raw = row.get(col, "")
                if raw == "" and col not in row:
                    continue                      # campo ausente != campo vacio
                value = clean(str(raw), col)
                cell = ws.cell(row=r, column=c)
                if infer_types:
                    cell.value = coerce(value)
                    if isinstance(cell.value, dt.datetime):
                        cell.number_format = "yyyy-mm-dd hh:mm:ss"
                else:
                    cell.value = value
                    cell.data_type = "s"          # nunca interpretar como formula
                if isinstance(cell.value, str) and cell.value.startswith(DANGEROUS):
                    cell.data_type = "s"

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(max(len(columns), 1))}{ws.max_row}"
        autosize(ws, columns, chunk)


def autosize(ws, columns, rows, sample: int = 200) -> None:
    for i, col in enumerate(columns, start=1):
        width = len(str(col))
        for row in rows[:sample]:
            width = max(width, len(str(row.get(col, ""))))
        ws.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 10), 60)


def sanitize_sheet_name(raw: str) -> str:
    name = re.sub(r"[\\/*?:\[\]]", "_", raw)
    if len(name) > 31:
        name = name[:14] + "~" + name[-16:]
    return name or "hoja"


def unique_sheet_name(wb: Workbook, raw: str) -> str:
    base = sanitize_sheet_name(raw)
    name, i = base, 2
    while name in wb.sheetnames:
        suffix = f"_{i}"
        name = base[:31 - len(suffix)] + suffix
        i += 1
    return name


def write_info_sheet(wb: Workbook, meta: dict) -> None:
    ws = wb.create_sheet(unique_sheet_name(wb, "_info"))
    ws.append(["Clave", "Valor"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for k, v in meta.items():
        ws.append([k, str(v)])
    if WARNINGS:
        ws.append([])
        ws.append(["Avisos", ""])
        for w in dict.fromkeys(WARNINGS):
            ws.append(["", w[:MAX_CELL_LEN]])
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 90
    for row in ws.iter_rows(min_col=2, max_col=2):
        row[0].alignment = Alignment(wrap_text=True, vertical="top")


# --------------------------------------------------------------------------
# Entrada interactiva de la ruta (cuando se ejecuta sin argumentos)
# --------------------------------------------------------------------------
# Opcional: si rellenas esta constante, se usara cuando no pases argumentos
# ni escribas nada en el prompt. Util para pulsar F5 en VS Code siempre sobre
# el mismo archivo. Dejala vacia para que siempre pregunte.
RUTA_POR_DEFECTO = ""


def normalize_path_input(raw: str) -> str:
    """Limpia lo que el usuario pega: comillas, espacios, ~, prefijos de shell."""
    s = (raw or "").strip()
    if s.startswith("& "):                 # PowerShell "Copiar como ruta de acceso"
        s = s[2:].strip()
    if len(s) >= 2 and s[0] == s[-1] and s[0] in "\"'":
        s = s[1:-1]
    s = s.replace("\\ ", " ")              # arrastrar y soltar en shells POSIX
    return str(Path(s.strip()).expanduser())


def pick_file_dialog() -> str | None:
    """Dialogo grafico de seleccion. Devuelve None si no hay entorno grafico."""
    try:
        import tkinter as tk
        from tkinter import filedialog
    except Exception:
        return None
    try:
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        path = filedialog.askopenfilename(
            title="Selecciona el archivo XML a convertir",
            filetypes=[("Archivos XML", "*.xml"), ("Todos los archivos", "*.*")],
        )
        root.destroy()
        return path or None
    except Exception:
        return None


def ask_xml_path(use_dialog: bool = True) -> str | None:
    """Pide la ruta del XML: dialogo grafico si se puede, si no por teclado."""
    if use_dialog:
        chosen = pick_file_dialog()
        if chosen:
            print(f"Archivo seleccionado: {chosen}")
            return chosen

    print("\n--- Conversor XML -> Excel ---")
    print("Arrastra el archivo a esta ventana o pega su ruta completa.")
    for intento in range(3):
        try:
            raw = input("Ruta del XML (Enter para cancelar): ")
        except (EOFError, KeyboardInterrupt):
            print()
            return None
        if not raw.strip():
            if RUTA_POR_DEFECTO:
                print(f"Usando la ruta por defecto: {RUTA_POR_DEFECTO}")
                return normalize_path_input(RUTA_POR_DEFECTO)
            return None
        path = normalize_path_input(raw)
        if Path(path).is_file():
            return path
        if Path(path).is_dir():
            print(f"  '{path}' es una carpeta, no un archivo.")
        else:
            print(f"  No se encuentra '{path}'. Revisa la ruta.")
        if intento == 2:
            print("Demasiados intentos.")
    return None


def ask_mode(default: str = "wide") -> str:
    opciones = {"1": "wide", "2": "relational", "3": "both"}
    print("\nModo de conversion:")
    print("  1) wide       - una fila por registro (por defecto)")
    print("  2) relational - listas anidadas en hojas separadas")
    print("  3) both       - ambas cosas")
    try:
        elegido = input("Opcion [1]: ").strip()
    except (EOFError, KeyboardInterrupt):
        return default
    return opciones.get(elegido, default)


# --------------------------------------------------------------------------
# Modo streaming (archivos muy grandes)
# --------------------------------------------------------------------------
def _stream_elements(path: str, tag: str, keep_ns: bool):
    nsmap: dict[str, str] = {}
    name = make_namer(keep_ns, nsmap)
    for event, payload in ET.iterparse(path, events=("start-ns", "end")):
        if event == "start-ns":
            prefix, uri = payload
            nsmap.setdefault(uri, prefix or "ns")
            continue
        elem = payload
        if isinstance(elem.tag, str) and name(elem.tag).split(":")[-1] == tag:
            yield elem, name
            elem.clear()


def stream_records(path: str, tag: str, keep_ns: bool):
    """Genera registros sin cargar todo el arbol en memoria (2 pasadas:
    la primera fija el esquema de columnas, la segunda emite las filas)."""
    schema = {"multi": set(), "mixed": set()}
    for elem, name in _stream_elements(path, tag, keep_ns):
        partial = build_schema([elem], name)
        schema["multi"] |= partial["multi"]
        schema["mixed"] |= partial["mixed"]
    for elem, name in _stream_elements(path, tag, keep_ns):
        yield flatten(elem, name, schema=schema)


# --------------------------------------------------------------------------
# main
# --------------------------------------------------------------------------
def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="Convierte XML a Excel sin perder campos.")
    ap.add_argument("xml", nargs="?", help="archivo XML de entrada (si se omite, se pregunta)")
    ap.add_argument("-o", "--output", help="archivo .xlsx de salida")
    ap.add_argument("--mode", choices=("wide", "relational", "both"), default=None)
    ap.add_argument("--no-dialog", action="store_true",
                    help="en modo interactivo, pedir la ruta por teclado sin ventana grafica")
    ap.add_argument("--record-path", help="ruta absoluta del elemento-registro, ej: root/items/item")
    ap.add_argument("--keep-namespaces", action="store_true", help="conserva prefijos ns en los nombres")
    ap.add_argument("--infer-types", action="store_true", help="convierte numeros/fechas/booleanos a tipo nativo")
    ap.add_argument("--list-paths", action="store_true", help="solo muestra las rutas detectadas y sale")
    ap.add_argument("--stream", metavar="TAG", help="modo streaming por nombre de etiqueta (XML enormes)")
    args = ap.parse_args(argv)

    global INTERACTIVO
    INTERACTIVO = args.xml is None
    if INTERACTIVO:
        elegido = ask_xml_path(use_dialog=not args.no_dialog)
        if not elegido:
            print("Operacion cancelada: no se indico ningun archivo XML.", file=sys.stderr)
            return 1
        args.xml = elegido
        if args.mode is None:
            args.mode = ask_mode()
    if args.mode is None:
        args.mode = "wide"

    src = Path(normalize_path_input(args.xml))
    if not src.exists():
        print(f"error: no existe {src}", file=sys.stderr)
        return 2
    if src.is_dir():
        print(f"error: {src} es una carpeta, no un archivo XML", file=sys.stderr)
        return 2
    out = Path(args.output) if args.output else src.with_suffix(".xlsx")

    wb = Workbook()
    wb.remove(wb.active)
    meta = {
        "Archivo origen": src.name,
        "Generado": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Modo": "stream" if args.stream else args.mode,
        "Tipos nativos": "si" if args.infer_types else "no (todo texto)",
        "Namespaces": "conservados" if args.keep_namespaces else "eliminados",
    }

    # ---------- streaming ----------
    if args.stream:
        rows, cols = [], []
        for flat in stream_records(str(src), args.stream, args.keep_namespaces):
            rows.append(flat)
            for k in flat:
                if k not in cols:
                    cols.append(k)
        if not rows:
            print(f"error: ninguna etiqueta <{args.stream}> encontrada", file=sys.stderr)
            return 3
        write_sheet(wb, args.stream, cols, rows, args.infer_types)
        meta["Registros"] = len(rows)
        meta["Columnas"] = len(cols)
        write_info_sheet(wb, meta)
        wb.save(out)
        print(f"OK: {out}  ({len(rows)} filas x {len(cols)} columnas)")
        return 0

    # ---------- carga completa ----------
    try:
        root, nsmap = parse_with_namespaces(str(src))
    except ET.ParseError as e:
        print(f"error: XML mal formado -> {e}", file=sys.stderr)
        return 2

    name = make_namer(args.keep_namespaces, nsmap)
    counts, depth = path_stats(root, name)

    if args.list_paths:
        for p, c in sorted(counts.items(), key=lambda kv: (depth[kv[0]], kv[0])):
            print(f"{c:>8} x  {p}")
        return 0

    record_path = args.record_path or detect_record_path(counts, depth)
    if record_path is None:
        record_path = name(root.tag)
        records = [root]
        meta["Deteccion"] = "sin repeticiones: el documento entero es 1 registro"
    else:
        records = select_by_path(root, name, record_path)
        if not records:
            print(f"error: la ruta '{record_path}' no coincide con ningun elemento. "
                  f"Usa --list-paths para ver las disponibles.", file=sys.stderr)
            return 3
        meta["Deteccion"] = "manual" if args.record_path else "automatica"

    meta["Elemento registro"] = record_path
    meta["Registros"] = len(records)

    schema = build_schema(records, name)

    if args.mode in ("wide", "both"):
        rows, cols = [], []
        for rec in records:
            flat = flatten(rec, name, schema=schema)
            rows.append(flat)
            for k in flat:
                if k not in cols:
                    cols.append(k)
        write_sheet(wb, record_path.rsplit("/", 1)[-1] or "datos", cols, rows, args.infer_types)
        meta["Columnas (wide)"] = len(cols)

    if args.mode in ("relational", "both"):
        tables = build_relational(records, name, record_path, schema)
        for path, (cols, rows) in tables.items():
            title = path.rsplit("/", 1)[-1]
            if args.mode == "both":
                title = f"rel_{title}"
            write_sheet(wb, title, cols, rows, args.infer_types)
        meta["Tablas (relacional)"] = ", ".join(p.rsplit("/", 1)[-1] for p in tables)

    write_info_sheet(wb, meta)
    wb.save(out)
    print(f"OK: {out}  ({len(records)} registros, hojas: {', '.join(wb.sheetnames)})")
    if WARNINGS:
        print(f"aviso: {len(set(WARNINGS))} incidencias registradas en la hoja _info")
    return 0


if __name__ == "__main__":
    codigo = main()
    if INTERACTIVO:
        # Evita que la ventana se cierre de golpe al ejecutar con doble clic.
        try:
            input("\nPulsa Enter para salir...")
        except (EOFError, KeyboardInterrupt):
            pass
    raise SystemExit(codigo)