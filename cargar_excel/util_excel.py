import zipfile

import pandas as pd
import win32com.client
import openpyxl
import os
from openpyxl import load_workbook
import time
import psutil
from openpyxl.utils import get_column_letter

def ocultar_hoja_excel(archivo_ruta: str, nombre_hoja_a_ocultar: str, modo_oculto: str = 'hidden'):
    """
    Oculta una hoja específica de un archivo Excel existente y guarda los cambios.

    Args:
        archivo_ruta (str): Ruta completa al archivo Excel.
        nombre_hoja_a_ocultar (str): Nombre exacto de la hoja que se quiere ocultar.
        modo_oculto (str): Modo de ocultación. Opciones: 
                           'hidden' (Ocultar normal, fácil de desocultar) o 
                           'very_hidden' (Muy oculto, requiere VBA para desocultar).
    """
    
    # Mapeo de modos de texto a valores numéricos de openpyxl
    MODOS = {
        'hidden': 'hidden',
        'very_hidden': 'veryHidden'
    }

    if modo_oculto.lower() not in MODOS:
        print(f"❌ Advertencia: Modo de ocultación '{modo_oculto}' no reconocido. Usando 'hidden'.")
        modo_oculto = 'hidden'

    try:
        # 1. Cargar el libro de trabajo (workbook)
        print(f"⚙️  Cargando el archivo: {archivo_ruta}...")
        libro = openpyxl.load_workbook(archivo_ruta)
        
        # 2. Seleccionar la hoja a ocultar
        if nombre_hoja_a_ocultar not in libro.sheetnames:
            print(f"❌ Error: La hoja '{nombre_hoja_a_ocultar}' no existe en el archivo.")
            return

        hoja = libro[nombre_hoja_a_ocultar]

        # 3. Ocultar la hoja estableciendo la propiedad 'sheet_state'
        # El valor se toma del mapeo MODOS
        hoja.sheet_state = MODOS[modo_oculto.lower()]

        # 4. Guardar el libro de trabajo con los cambios
        libro.save(archivo_ruta)
        
        print("\n" + "-" * 60)
        print("🎉 Éxito al ocultar hoja:")
        print(f"   La hoja '{nombre_hoja_a_ocultar}' ha sido configurada como '{modo_oculto.upper()}'.")
        print(f"   Archivo modificado: {os.path.abspath(archivo_ruta)}")
        print("-" * 60)

    except FileNotFoundError:
        print(f"❌ Error: Archivo no encontrado en la ruta: {archivo_ruta}")
    except Exception as e:
        print(f"❌ Ocurrió un error: {e}")
        print("   Asegúrese de que el archivo no esté abierto por otra aplicación.")

# ## --- Ejemplo de Uso de la Función ---

# if __name__ == '__main__':
#     # ⚠️ REQUISITO: pip install openpyxl
    
#     ARCHIVO_MAESTRO = 'Reporte_Final_Seguridad.xlsx'
#     HOJA_A_ESCONDER = 'Datos_Sensibles'
    
#     print("Iniciando la ocultación de hoja...")
    
#     # Ocultación simple ('hidden'): el usuario puede desocultar desde el menú de Excel.
#     ocultar_hoja_excel(
#         archivo_ruta=ARCHIVO_MAESTRO,
#         nombre_hoja_a_ocultar=HOJA_A_ESCONDER,
#         modo_oculto='hidden'
#     )
    
#     # Ocultación avanzada ('very_hidden'): requiere código VBA para desocultar.
#     # ocultar_hoja_excel(
#     #     archivo_ruta=ARCHIVO_MAESTRO,
#     #     nombre_hoja_a_ocultar=HOJA_A_ESCONDER,
#     #     modo_oculto='very_hidden'
#     # )

def cerrar_excel_abierto(ruta_archivo: str):
    """
    Cierra un archivo Excel abierto en la aplicación Excel (si lo está),
    guardando los cambios antes de cerrarlo.
    
    Requiere: pywin32 (instalar con `pip install pywin32`)
    
    Parámetros:
        ruta_archivo (str): Ruta completa del archivo Excel a cerrar.
    """

    ruta_absoluta = os.path.abspath(ruta_archivo)

    # Conectarse a la aplicación Excel (si está abierta)
    try:
        excel = win32com.client.GetActiveObject("Excel.Application")
    except Exception:
        print("⚠️ No hay instancias activas de Excel abiertas.")
        return

    # Recorrer los libros abiertos
    for wb in excel.Workbooks:
        if os.path.abspath(wb.FullName).lower() == ruta_absoluta.lower():
            print(f"🧩 Guardando y cerrando: {wb.Name}")
            #######wb.Save()       # Guardar cambios
            wb.Close(SaveChanges=True)  # Cerrar archivo
            print(f"✅ Archivo cerrado correctamente: {ruta_archivo}")
            break
    else:
        print(f"⚠️ El archivo '{ruta_archivo}' no estaba abierto en Excel.")

    # Si no quedan libros abiertos, opcionalmente cerrar Excel completo
    if excel.Workbooks.Count == 0:
        excel.Quit()
        print("🛑 Excel se ha cerrado completamente.")

    # Liberar recursos COM
    del excel

# def inmovilizar_fila(ruta_archivo: str, nombre_hoja: str,freeze_panes_row: str="A2"):
#     """
#     Inmoviliza (congela) la primera fila de una hoja específica en un archivo Excel existente.
    
#     Parámetros:
#         ruta_archivo (str): Ruta del archivo Excel.
#         nombre_hoja (str): Nombre de la hoja en la cual se quiere inmovilizar la primera fila.
#     """

#     # Cargar el archivo Excel
#     wb = load_workbook(ruta_archivo)

#     # Verificar que la hoja exista
#     if nombre_hoja not in wb.sheetnames:
#         raise ValueError(f"La hoja '{nombre_hoja}' no existe en el archivo '{ruta_archivo}'.")

#     # Seleccionar la hoja
#     ws = wb[nombre_hoja]

#     # Congelar la primera fila
#     # Esto deja visibles las celdas hasta la fila 1 cuando se hace scroll
#     ws.freeze_panes = freeze_panes_row #"A2"

#     # Guardar los cambios
#     wb.save(ruta_archivo)
#     print(f"✅ Se inmovilizó la fila {freeze_panes_row} de la hoja '{nombre_hoja}'.")

def inmovilizar_fila(ruta_archivo: str, nombre_hoja: str, freeze_panes_row: str = "A2"):
    """
    Inmoviliza (congela) la primera fila de una hoja específica en un archivo Excel existente.
    """
    try:
        # 1. Validar que el archivo exista y no esté vacío antes de intentar abrirlo
        if not os.path.exists(ruta_archivo) or os.path.getsize(ruta_archivo) == 0:
            print(f"⚠️ El archivo '{ruta_archivo}' no existe o está vacío (0 KB).")
            return

        # 2. Intentar cargar el archivo Excel
        wb = load_workbook(ruta_archivo)

        # Verificar que la hoja exista
        if nombre_hoja not in wb.sheetnames:
            wb.close() # Cerrar el flujo antes de lanzar el error
            raise ValueError(f"La hoja '{nombre_hoja}' no existe en el archivo '{ruta_archivo}'.")

        # Seleccionar la hoja e inmovilizar
        ws = wb[nombre_hoja]
        ws.freeze_panes = freeze_panes_row

        # Guardar y cerrar de forma segura
        wb.save(ruta_archivo)
        wb.close()
        print(f"✅ Se inmovilizó la fila {freeze_panes_row} de la hoja '{nombre_hoja}'.")

    except zipfile.BadZipFile:
        print(f"❌ Error crítico: '{ruta_archivo}' no es un archivo zip válido (está corrupto).")
        # Opcional: Eliminar el archivo dañado para que no cause problemas en la siguiente ejecución
        try:
            os.remove(ruta_archivo)
            print(f"🧹 Archivo corrupto eliminado para limpieza.")
        except Exception:
            pass

    except PermissionError:
        print(f"❌ Error de permisos: El archivo '{ruta_archivo}' está abierto en Microsoft Excel. Ciérralo.")
        
    except Exception as e:
        print(f"❌ Ocurrió un error inesperado: {e}")


def eliminar_archivo_excel(ruta_archivo: str, reintentos: int = 3):
    """
    Cierra y elimina un archivo Excel específico liberando los recursos.
    Guarda los cambios antes de cerrarlo si está abierto en Excel.

    ⚙️ Requisitos:
        pip install pywin32 psutil

    Parámetros:
        ruta_archivo (str): Ruta completa del archivo Excel a eliminar.
        reintentos (int): Número de intentos de eliminación si el archivo está bloqueado.
    """

    ruta_absoluta = os.path.abspath(ruta_archivo)

    # === Paso 1: Intentar cerrar el archivo si está abierto en Excel ===
    try:
        excel = win32com.client.GetActiveObject("Excel.Application")
        for wb in excel.Workbooks:
            if os.path.abspath(wb.FullName).lower() == ruta_absoluta.lower():
                print(f"🧩 Guardando y cerrando: {wb.Name}")
                wb.Save()
                wb.Close(SaveChanges=True)
                print("✅ Archivo cerrado correctamente desde Excel.")
                break
        del excel
    except Exception:
        pass  # No había instancia activa de Excel

    # === Paso 2: Esperar a que se liberen los recursos ===
    for intento in range(reintentos):
        try:
            os.remove(ruta_absoluta)
            print(f"🗑️ Archivo eliminado correctamente: {ruta_absoluta}")
            return
        except PermissionError:
            print(f"⚠️ El archivo está en uso, reintentando ({intento+1}/{reintentos})...")
            time.sleep(1)
        except FileNotFoundError:
            print("❌ El archivo no existe o ya fue eliminado.")
            return
        except Exception as e:
            print(f"❌ Error inesperado: {e}")
            return

    # === Paso 3: Forzar liberación si sigue bloqueado ===
    print("⚠️ No se pudo eliminar después de varios intentos. Verificando procesos Excel...")

    for proceso in psutil.process_iter(['name']):
        if proceso.info['name'] and 'EXCEL.EXE' in proceso.info['name'].upper():
            proceso.terminate()
            print("🛑 Proceso de Excel cerrado a la fuerza.")
            time.sleep(1)

    try:
        os.remove(ruta_absoluta)
        print(f"✅ Archivo eliminado tras liberar proceso: {ruta_absoluta}")
    except Exception as e:
        print(f"❌ No se pudo eliminar el archivo: {e}")

def establecer_ancho_columnas(
    archivo_ruta: str, 
    hoja_nombre: str, 
    columnas_ancho: dict
):
    """
    Establece un ancho específico a un conjunto de columnas de una hoja de Excel 
    y sobrescribe esa hoja, manteniendo el resto de hojas intactas.

    Args:
        archivo_ruta (str): Ruta del archivo Excel existente.
        hoja_nombre (str): Nombre de la hoja a modificar.
        columnas_ancho (dict): Diccionario donde la clave es el nombre de la columna 
                               y el valor es el ancho (e.g., {'Nombre': 20, 'ID': 10}).
    """
    
    try:
        # 1. Cargar TODAS las hojas del archivo
        print(f"⚙️  Cargando todas las hojas del archivo '{archivo_ruta}'...")
        xls = pd.ExcelFile(archivo_ruta)
        # Diccionario con todas las hojas: {'NombreHoja': DataFrame}
        diccionario_hojas = pd.read_excel(xls, sheet_name=xls.sheet_names)
        
        if hoja_nombre not in diccionario_hojas:
            print(f"❌ Error: La hoja '{hoja_nombre}' no se encontró en el archivo.")
            return

        df_a_modificar = diccionario_hojas[hoja_nombre]
        print(f"✅ Hoja '{hoja_nombre}' cargada. Total de filas: {len(df_a_modificar)}")
        
    except FileNotFoundError:
        print(f"❌ Error: Archivo no encontrado en la ruta: {archivo_ruta}")
        return
    except Exception as e:
        print(f"❌ Error al cargar el archivo/hojas: {e}")
        return

    # 2. Guardar TODAS las hojas de nuevo (Sobreescritura con preservación)
    print(f"⚙️  Aplicando ancho a la hoja '{hoja_nombre}' y reescribiendo el archivo...")
    
    try:
        # Usamos mode='w' (write/escribir) con engine='xlsxwriter' (necesario para formatos)
        with pd.ExcelWriter(archivo_ruta, engine='xlsxwriter', mode='w') as writer:
            
            # Recorrer todas las hojas originales
            for nombre_hoja, df in diccionario_hojas.items():
                
                # Escribir el DataFrame en la hoja
                df.to_excel(writer, sheet_name=nombre_hoja, index=False)
                
                # 3. Aplicar ancho de columna SOLO si es la hoja especificada
                if nombre_hoja == hoja_nombre:
                    # Acceder al objeto Worksheet de xlsxwriter
                    worksheet = writer.sheets[nombre_hoja]
                    
                    # Mapear los nombres de columna a sus índices (0, 1, 2, ...)
                    columnas = df.columns.tolist()
                    
                    for nombre_columna, ancho in columnas_ancho.items():
                        try:
                            # Obtener el índice (número) de la columna
                            col_indice = columnas.index(nombre_columna)
                            
                            # Establecer el ancho de la columna (col_start, col_end, width)
                            # Se usa set_column(col_indice, col_indice, ancho)
                            worksheet.set_column(col_indice, col_indice, ancho)
                            print(f"   -> Ancho {ancho} aplicado a columna '{nombre_columna}'.")
                        except ValueError:
                            print(f"   ⚠️ Advertencia: Columna '{nombre_columna}' no encontrada en la hoja.")
                            
        print("\n" + "-" * 60)
        print("🎉 Éxito en la operación:")
        print(f"   El ancho de las columnas en la hoja '{hoja_nombre}' ha sido actualizado.")
        print(f"   El resto de hojas se mantienen intactas.")
        print(f"   Archivo modificado: {os.path.abspath(archivo_ruta)}")
        print("-" * 60)

    except Exception as e:
        print(f"❌ Error al escribir en el archivo: {e}")

## --- Ejemplo de Uso de la Función ---

# if __name__ == '__main__':
#     # Define tus parámetros.
#     ARCHIVO_MAESTRO = 'Reporte_Consolidado.xlsx'
#     HOJA_A_MODIFICAR = 'Datos_Principales'
    
#     # Diccionario: {'Nombre exacto de la columna': Ancho deseado}
#     ANCHO_COLUMNAS = {
#         'Clave_Unificada': 30,
#         'Nombre Completo': 45,
#         'ID_SistemaA': 15,
#         'Fecha Creacion': 18
#     }
    
#     print("Iniciando el proceso de ajuste de ancho de columnas...")
    
#     # Requisitos: pip install pandas openpyxl xlsxwriter
    
#     establecer_ancho_columnas(
#         archivo_ruta=ARCHIVO_MAESTRO,
#         hoja_nombre=HOJA_A_MODIFICAR,
#         columnas_ancho=ANCHO_COLUMNAS
#     )

def establecer_ancho_columnas_excel(
    archivo_ruta: str, 
    hoja_nombre: str, 
    columnas_ancho: dict
):
    """
    Establece un ancho específico a un conjunto de columnas de una hoja de Excel, 
    preservando el resto de hojas y todas las configuraciones previas (filtros, colores).

    Args:
        archivo_ruta (str): Ruta del archivo Excel existente.
        hoja_nombre (str): Nombre de la hoja a modificar.
        columnas_ancho (dict): Diccionario donde la clave es el nombre EXACTO de la columna 
                               y el valor es el ancho deseado (e.g., {'Nombre': 20, 'ID': 10}).
    """
    
    try:
        # 1. Cargar el Workbook existente (openpyxl preserva la estructura del archivo)
        print(f"⚙️  Cargando archivo: '{archivo_ruta}'...")
        workbook = openpyxl.load_workbook(archivo_ruta)

        # 2. Seleccionar la hoja específica
        if hoja_nombre not in workbook.sheetnames:
            print(f"❌ Error: La hoja '{hoja_nombre}' no se encontró en el archivo.")
            return

        worksheet = workbook[hoja_nombre]
        print(f"✅ Hoja '{hoja_nombre}' cargada.")

        # 3. Mapear nombres de columna a letras de columna (A, B, C, ...)
        # Se asume que la primera fila (row 1) contiene los encabezados.
        # openpyxl usa un índice basado en 1 (fila 1, columna 1)
        encabezados = [cell.value for cell in worksheet[1]]
        
        columna_mapa = {}
        for idx, header in enumerate(encabezados):
            if header in columnas_ancho:
                # get_column_letter convierte el índice basado en 1 (idx + 1) a la letra (A, B, ...)
                col_letra = get_column_letter(idx + 1)
                columna_mapa[header] = col_letra
        
        # 4. Aplicar el ancho de columna
        print("⚙️  Estableciendo ancho de columnas...")
        
        columnas_modificadas = 0
        for nombre_columna, ancho in columnas_ancho.items():
            if nombre_columna in columna_mapa:
                col_letra = columna_mapa[nombre_columna]
                
                # Modificar la dimensión de la columna para establecer el ancho
                worksheet.column_dimensions[col_letra].width = ancho
                columnas_modificadas += 1
                print(f"   -> Ancho {ancho} aplicado a columna '{nombre_columna}' ({col_letra}).")
            else:
                print(f"   ⚠️ Advertencia: Columna '{nombre_columna}' no encontrada en la hoja.")
        
        if columnas_modificadas == 0:
            print("⚠️ Advertencia: Ninguna de las columnas especificadas fue encontrada y modificada.")
            return

        # 5. Guardar el archivo (sobrescribe el existente)
        # openpyxl guarda los cambios realizados en el objeto workbook, preservando el resto.
        workbook.save(archivo_ruta)
        
        print("\n" + "-" * 60)
        print("🎉 Éxito en la operación:")
        print(f"   El ancho de las columnas en '{hoja_nombre}' ha sido ajustado.")
        print(f"   **Filtros, colores y otras hojas se mantienen intactos.**")
        print(f"   Archivo modificado: {os.path.abspath(archivo_ruta)}")
        print("-" * 60)

    except FileNotFoundError:
        print(f"❌ Error: Archivo no encontrado en la ruta: {archivo_ruta}")
    except Exception as e:
        print(f"❌ Ocurrió un error inesperado: {e}")

## --- Ejemplo de Uso de la Función ---

# if __name__ == '__main__':
#     # Define tus parámetros.
#     ARCHIVO_MAESTRO = 'Reporte_De_Ventas.xlsx'
#     HOJA_A_MODIFICAR = 'Datos_Transacciones'
    
#     # Diccionario: {'Nombre exacto de la columna': Ancho deseado}
#     ANCHO_COLUMNAS = {
#         'ID_Transaccion': 15,
#         'Descripción del Producto': 50,
#         'Fecha de Venta': 18,
#         'Monto Total': 15
#     }
    
#     print("Iniciando el proceso de ajuste de ancho de columnas...")
    
#     # ⚠️ Requisito:
#     # pip install openpyxl
    
#     establecer_ancho_columnas_excel(
#         archivo_ruta=ARCHIVO_MAESTRO,
#         hoja_nombre=HOJA_A_MODIFICAR,
#         columnas_ancho=ANCHO_COLUMNAS
#     )