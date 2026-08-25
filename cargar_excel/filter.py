import os
import time

import pandas as pd
from openpyxl import load_workbook

def copiar_registros_filtrados(
    ruta_archivo: str,
    hoja_origen: str,
    nombre_columna: str = "Estado",
    valor_columna: str = "deshabilitado",
    hoja_destino: str = "Registros_Deshabilitados"
):
    """
    Filtra los registros de una hoja de Excel donde una columna tenga el valor 'deshabilitado'
    y los copia en una nueva hoja dentro del mismo archivo.
    Compatible con pandas >= 2.2.
    """

    # === Leer la hoja de origen ===
    df = pd.read_excel(ruta_archivo, sheet_name=hoja_origen, engine='openpyxl')

    # === Validar la existencia de la columna ===
    if nombre_columna not in df.columns:
        raise ValueError(f"La columna '{nombre_columna}' no existe en la hoja '{hoja_origen}'.")

    # === Filtrar los registros deshabilitados ===
    df_filtrado = df[df[nombre_columna].astype(str).str.lower() == valor_columna.lower()]

    if df_filtrado.empty:
        print(f"⚠️ No se encontró en la columna {nombre_columna} registros con estado {valor_columna}.")
        return

    # === Cargar el libro existente ===
    book = load_workbook(ruta_archivo)

    # Si la hoja destino ya existe, eliminarla
    if hoja_destino in book.sheetnames:
        del book[hoja_destino]

    # === Crear el escritor Excel vinculado al libro ===
    with pd.ExcelWriter(ruta_archivo, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        writer._book = book  # se usa _book en lugar de book
        df_filtrado.to_excel(writer, sheet_name=hoja_destino, index=False)

    print(f"✅ Se copiaron {len(df_filtrado)} registros deshabilitados a la hoja '{hoja_destino}'.")


def mostrar_columnas_y_registros(
    archivo_ruta: str, 
    hoja_nombre: str, 
    columnas_a_mostrar: list, 
    n_registros: int = 5
):
    """
    Carga una hoja de Excel, selecciona solo las columnas indicadas e imprime 
    los primeros N registros.

    Args:
        archivo_ruta (str): Ruta del archivo Excel.
        hoja_nombre (str): Nombre de la hoja a leer.
        columnas_a_mostrar (list): Lista de strings con los nombres exactos de las columnas a incluir.
        n_registros (int): Número de registros a mostrar (por defecto 5).
    """
    try:
        # 1. Cargar el DataFrame desde la hoja, seleccionando SOLO las columnas deseadas
        # El parámetro 'usecols' es clave para cargar un subconjunto de columnas.
        print(f"⚙️  Cargando hoja '{hoja_nombre}' y columnas específicas...")
        df = pd.read_excel(
            archivo_ruta, 
            sheet_name=hoja_nombre, 
            usecols=columnas_a_mostrar
        )
        
        # 2. Usar el método .head() para obtener los primeros N registros
        registros_seleccionados = df.head(n_registros)
        
        print(f"\n✅ Primeros {n_registros} registros de las columnas seleccionadas:")
        print("-" * 70)
        
        # 3. Imprimir el resultado
        print(registros_seleccionados)
        
        print("-" * 70)
        print(f"Columnas mostradas: {columnas_a_mostrar}")
        print(f"Total de registros cargados (para esta muestra): {len(df)}")

    except FileNotFoundError:
        print(f"❌ Error: Archivo no encontrado en la ruta: {archivo_ruta}")
    except KeyError:
        print(f"❌ Error: Una o más de las columnas o la hoja no se encontró en el archivo.")
    except Exception as e:
        print(f"❌ Ocurrió un error: {e}")

def aplicar_filtros_excel(ruta_archivo: str, hoja_nombre: str, fila_encabezado: int = 1):
    """
    Aplica filtros automáticos a una hoja de un archivo Excel existente.
    Detecta automáticamente la cantidad de columnas (campos dinámicos).

    Parámetros:
        ruta_archivo (str): Ruta del archivo Excel.
        hoja_nombre (str): Nombre de la hoja donde se aplicarán los filtros.
        fila_encabezado (int): Número de la fila que contiene los encabezados (por defecto, 1).
    """
    try:
        # === Cargar archivo ===
        wb = load_workbook(ruta_archivo)
        if hoja_nombre not in wb.sheetnames:
            raise ValueError(f"La hoja '{hoja_nombre}' no existe en el archivo.")
        
        ws = wb[hoja_nombre]

        # === Detectar última columna con datos ===
        ultima_columna = ws.max_column
        ultima_fila = ws.max_row

        # Convertir número de columna a letra (por ejemplo: 5 → 'E')
        from openpyxl.utils import get_column_letter
        letra_ultima_columna = get_column_letter(ultima_columna)

        # === Definir rango del autofiltro ===
        rango = f"A{fila_encabezado}:{letra_ultima_columna}{ultima_fila}"

        # Aplicar autofiltro
        ws.auto_filter.ref = rango

        # === Guardar cambios ===
        wb.save(ruta_archivo)
        print(f"✅ Filtros aplicados correctamente en la hoja '{hoja_nombre}' (rango {rango}).")
    except Exception as error:
                print(f"Error en método aplicar_filtros_excel '{hoja_nombre}' (rango {rango}): {error}")

import zipfile
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def aplicar_filtros_excel_validado(ruta_archivo: str, hoja_nombre: str, fila_encabezado: int = 1):
    """
    Aplica filtros automáticos a una hoja de un archivo Excel existente.
    Detecta automáticamente la cantidad de columnas (campos dinámicos).
    """
    # Inicializamos la variable en None para evitar el error de acceso local
    rango = None 
    
    try:
        # === Cargar archivo ===
        wb = load_workbook(ruta_archivo)
        if hoja_nombre not in wb.sheetnames:
            wb.close()
            raise ValueError(f"La hoja '{hoja_nombre}' no existe en el archivo.")
        
        ws = wb[hoja_nombre]

        # === Detectar dimensiones ===
        ultima_columna = ws.max_column
        ultima_fila = ws.max_row
        letra_ultima_columna = get_column_letter(ultima_columna)

        # === Definir rango del autofiltro ===
        rango = f"A{fila_encabezado}:{letra_ultima_columna}{ultima_fila}"

        # Aplicar autofiltro
        ws.auto_filter.ref = rango

        # === Guardar y cerrar cambios ===
        wb.save(ruta_archivo)
        wb.close()
        print(f"✅ Filtros aplicados correctamente en la hoja '{hoja_nombre}' (rango {rango}).")
        time.sleep(2) 
    except zipfile.BadZipFile:
        print(f"❌ Error en '{hoja_nombre}': El archivo Excel está corrupto o vacío (BadZipFile).")
        
    except Exception as error:
        # Si el error ocurrió antes de definir el rango, mostramos un mensaje alternativo
        if rango:
            print(f"❌ Error en método aplicar_filtros_excel '{hoja_nombre}' (rango {rango}): {error}")
        else:
            print(f"❌ Error en método aplicar_filtros_excel '{hoja_nombre}' antes de definir el rango: {error}")

def inmovilizar_filas_por_hoja(ruta_archivo: str, configuracion_hojas: dict):
    """
    Inmoviliza filas en múltiples hojas de un archivo Excel abriéndolo una sola vez.
    
    Parámetros:
        ruta_archivo (str): Ruta o nombre del archivo Excel.
        configuracion_hojas (dict): Diccionario donde la CLAVE es el nombre de la hoja 
                                    y el VALOR es la celda de corte (ej: "A2" para fila 1, "A3" para fila 2).
    """
    try:
        # 1. Validar que el archivo exista y tenga contenido
        if not os.path.exists(ruta_archivo) or os.path.getsize(ruta_archivo) == 0:
            print(f"❌ El archivo '{ruta_archivo}' no existe o está vacío.")
            return

        # 2. Abrir el archivo una sola vez en memoria
        print(f"Buscando archivo: {ruta_archivo}...")
        wb = load_workbook(ruta_archivo)
        
        # 3. Recorrer la configuración hoja por hoja
        for nombre_hoja, celda_inmovilizar in configuracion_hojas.items():
            if nombre_hoja in wb.sheetnames:
                ws = wb[nombre_hoja]
                
                # Aplicar la inmovilización (ej: "A2" congela la fila 1)
                ws.freeze_panes = celda_inmovilizar
                print(f"  ✅ Fila configurada en '{nombre_hoja}' usando corte en {celda_inmovilizar}.")
            else:
                print(f"  ⚠️ Advertencia: La hoja '{nombre_hoja}' no existe en este archivo. Saltando...")

        # 4. Guardar y cerrar el archivo de forma segura
        wb.save(ruta_archivo)
        wb.close()
        print(f"🎉 ¡Cambios guardados con éxito en '{ruta_archivo}'!")

    except zipfile.BadZipFile:
        print(f"❌ Error crítico: El archivo '{ruta_archivo}' está corrupto o incompleto (BadZipFile).")
    except PermissionError:
        print(f"❌ Error de permisos: El archivo '{ruta_archivo}' está abierto en Excel. Ciérralo antes de ejecutar.")
    except Exception as e:
        print(f"❌ Ocurrió un error inesperado: {e}")
        
import os
import zipfile
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def aplicar_filtros_e_inmovilizar_lote(ruta_archivo: str, configuracion_hojas: dict, fila_encabezado: int = 1):
    """
    Optimiza recursos abriendo el archivo Excel UNA SOLA VEZ.
    Aplica autofiltros dinámicos e inmoviliza la fila de encabezado en todas las hojas dadas.

    Parámetros:
        ruta_archivo (str): Ruta del archivo Excel (.xlsx).
        configuracion_hojas (dict): Diccionario donde la clave es el nombre de la hoja 
                                    y el valor es la celda para inmovilizar (ej: "A2").
        fila_encabezado (int): Línea donde empiezan los filtros (por defecto, 1).
    """
    try:
        # === 1. Validación previa del archivo ===
        if not os.path.exists(ruta_archivo) or os.path.getsize(ruta_archivo) == 0:
            print(f"❌ El archivo '{ruta_archivo}' no existe o está vacío (0 KB).")
            return

        # === 2. Cargar el libro de trabajo una sola vez ===
        print(f"📂 Abriendo archivo para procesamiento por lote: {ruta_archivo}")
        wb = load_workbook(ruta_archivo)
        
        # === 3. Procesar cada hoja en memoria ===
        for hoja_nombre, celda_inmovilizar in configuracion_hojas.items():
            if hoja_nombre not in wb.sheetnames:
                print(f"⚠️ Advertencia: La hoja '{hoja_nombre}' no existe en el archivo. Saltando...")
                continue
            
            ws = wb[hoja_nombre]

            # Detectar dimensiones dinámicas de la hoja
            ultima_columna = ws.max_column
            ultima_fila = ws.max_row
            
            # Si la hoja está completamente vacía, evitar aplicar filtros erróneos
            if ultima_columna == 0 or ultima_fila == 0:
                print(f"⚠️ La hoja '{hoja_nombre}' está vacía. No se aplicaron filtros.")
                continue

            letra_ultima_columna = get_column_letter(ultima_columna)

            # Definir rango y aplicar el autofiltro
            rango_filtro = f"A{fila_encabezado}:{letra_ultima_columna}{ultima_fila}"
            ws.auto_filter.ref = rango_filtro

            # Inmovilizar la fila de encabezado usando la celda provista
            if celda_inmovilizar:
                ws.freeze_panes = celda_inmovilizar

            print(f"  ✅ Filtros ({rango_filtro}) e inmovilización ({celda_inmovilizar}) listos en '{hoja_nombre}'.")

        # === 4. Guardar cambios y liberar el archivo una sola vez ===
        wb.save(ruta_archivo)
        wb.close()
        print(f"🎉 ¡Procesamiento completado con éxito! Archivo guardado y cerrado de forma segura.")

    except zipfile.BadZipFile:
        print(f"❌ Error crítico: El archivo '{ruta_archivo}' está corrupto o incompleto (BadZipFile).")
    except PermissionError:
        print(f"❌ Error de permisos: El archivo '{ruta_archivo}' está abierto en Excel. Ciérralo.")
    except Exception as error:
        print(f"❌ Ocurrió un error inesperado durante el procesamiento por lote: {error}")

def copiar_registros_filtrados_lote(ruta_archivo: str, hoja_origen: str, operaciones_filtrado: list):
    """
    Optimiza recursos abriendo el archivo una sola vez.
    Aplica múltiples filtros en memoria sobre una hoja de origen y los guarda en pestañas separadas.

    Parámetros:
        ruta_archivo (str): Ruta del archivo Excel (.xlsx).
        hoja_origen (str): Nombre de la hoja base desde donde se leerán los datos.
        operaciones_filtrado (list): Lista de diccionarios con la configuración de cada filtro.
    """
    if not os.path.exists(ruta_archivo) or os.path.getsize(ruta_archivo) == 0:
        print(f"❌ Error: El archivo '{ruta_archivo}' no existe o está vacío.")
        return

    try:
        print(f"📂 Cargando hoja de origen '{hoja_origen}' desde: {ruta_archivo}")
        # 1. Leer la hoja origen una sola vez
        df_origen = pd.read_excel(ruta_archivo, sheet_name=hoja_origen, engine='openpyxl')
        
        # Diccionario para almacenar los DataFrames filtrados listos para guardar
        resultados_a_escribir = {}

        # 2. Procesar cada filtro directamente en la RAM
        for op in operaciones_filtrado:
            columna = op['nombre_columna']
            valor = str(op['valor_columna']).lower()
            destino = op['hoja_destino']

            # Validar existencia de la columna
            if columna not in df_origen.columns:
                print(f"  ❌ Error: La columna '{columna}' no existe en '{hoja_origen}'. Saltando este filtro.")
                continue

            # Filtrar los registros comparando de forma segura como texto en minúsculas
            df_filtrado = df_origen[df_origen[columna].astype(str).str.lower() == valor]

            if df_filtrado.empty:
                print(f"  ⚠️ Advertencia: No se encontraron registros con '{columna}' = '{valor}' para la hoja '{destino}'.")
                # Guardamos un DataFrame vacío con los mismos encabezados para no romper estructuras
                resultados_a_escribir[destino] = pd.DataFrame(columns=df_origen.columns)
            else:
                resultados_a_escribir[destino] = df_filtrado
                print(f"  🧠 Filtrado listo en RAM: {len(df_filtrado)} registros preparados para '{destino}'.")

        # 3. Guardar todas las hojas generadas en un solo bloque seguro de escritura
        if resultados_a_escribir:
            print(f"💾 Escribiendo todas las nuevas hojas filtradas en el disco...")
            with pd.ExcelWriter(ruta_archivo, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                for hoja_destino, df_datos in resultados_a_escribir.items():
                    df_datos.to_excel(writer, sheet_name=hoja_destino, index=False)
            print(f"🎉 ¡Todas las hojas filtradas se crearon con éxito en '{ruta_archivo}'!")
        else:
            print("⚠️ No se realizó ninguna escritura porque no hubo filtros válidos.")

    except zipfile.BadZipFile:
        print(f"❌ Error crítico: El archivo '{ruta_archivo}' está corrupto (BadZipFile).")
    except PermissionError:
        print(f"❌ Error de permisos: El archivo '{ruta_archivo}' está abierto en Excel. Ciérralo.")
    except Exception as error:
        print(f"❌ Ocurrió un error inesperado al filtrar el lote: {error}")


