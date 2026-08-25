import zipfile

from openpyxl import load_workbook
import pandas as pd
import os

def mezclar_excel(
    archivo1_ruta: str, 
    archivo2_ruta: str, 
    hoja1_nombre: str, 
    hoja2_nombre: str, 
    clave_archivo1: str, 
    clave_archivo2: str, 
    nombre_nuevo_archivo: str,
    tipo_join: str = 'outer'
):
    """
    Realiza la mezcla (merge) de dos archivos Excel usando columnas clave con 
    nombres distintos, sin renombrar dichas columnas.

    Args:
        archivo1_ruta (str): Ruta del primer archivo Excel.
        archivo2_ruta (str): Ruta del segundo archivo Excel.
        hoja1_nombre (str): Nombre de la hoja a usar en el primer archivo.
        hoja2_nombre (str): Nombre de la hoja a usar en el segundo archivo.
        clave_archivo1 (str): Nombre EXACTO de la columna clave en el primer archivo.
        clave_archivo2 (str): Nombre EXACTO de la columna clave en el segundo archivo.
        nombre_nuevo_archivo (str): Nombre del nuevo archivo Excel de salida.
        tipo_join (str): Tipo de join a usar ('inner', 'outer', 'left', 'right'). Por defecto 'outer'.
    """
    
    try:
        # 1. Cargar los archivos y hojas específicas
        print(f"⚙️  Cargando hojas: '{hoja1_nombre}' y '{hoja2_nombre}'...")
        df1 = pd.read_excel(archivo1_ruta, sheet_name=hoja1_nombre)
        df2 = pd.read_excel(archivo2_ruta, sheet_name=hoja2_nombre)

        # 2. Realizar la mezcla (Merge) sin renombrar
        # Se usa 'left_on' para la clave del df1 y 'right_on' para la clave del df2.
        # Esto permite hacer el match aunque los nombres de las columnas sean distintos.
        df_mezclado = pd.merge(
            df1, 
            df2, 
            left_on=clave_archivo1,  # Columna clave en el primer DataFrame
            right_on=clave_archivo2, # Columna clave en el segundo DataFrame
            how=tipo_join, 
            suffixes=('_SistemaA', '_SistemaB')
        )

        #print(f"✅ Mezcla completada. Filas totales: {len(df_mezclado)}")

        # 3. Guardar el DataFrame mezclado en un nuevo archivo
        df_mezclado.to_excel(nombre_nuevo_archivo,sheet_name='MergeUsr', index=False)

        print("\n" + "-" * 60)
        print("🎉 Éxito en el procesamiento:")
        print(f"   El resultado se ha guardado en el archivo: {os.path.abspath(nombre_nuevo_archivo)} tipo: {tipo_join}")
        print("-" * 60)

    except FileNotFoundError:
        print("❌ Error: Uno o ambos archivos de entrada no se encontraron.")
    except KeyError as e:
        print(f"❌ Error: El nombre de la columna o de la hoja no se encontró. Verifique la sintaxis.")
        print(f"         El parámetro faltante: {e}")
    except Exception as e:
        print(f"❌ Ocurrió un error inesperado durante el proceso: {e}")

def unir_hojas_excel(
    ruta_archivo: str,
    hoja_1: str,
    hoja_2: str,
    campo_hoja_1: str,
    campo_hoja_2: str,
    hoja_destino: str = "Hoja_Unida",
    tipo_union: str = "inner"
):
    """
    Une dos hojas específicas de un archivo Excel en una nueva hoja, 
    considerando que los campos de unión pueden tener nombres distintos.
    El archivo original se conserva y solo se agrega una nueva hoja con el resultado.

    Parámetros:
        ruta_archivo (str): Ruta del archivo Excel existente.
        hoja_1 (str): Nombre de la primera hoja a unir.
        hoja_2 (str): Nombre de la segunda hoja a unir.
        campo_hoja_1 (str): Nombre de la columna en la hoja 1 usada para unir.
        campo_hoja_2 (str): Nombre de la columna en la hoja 2 usada para unir.
        hoja_destino (str): Nombre de la hoja donde se guardará el resultado.
        tipo_union (str): Tipo de unión ('inner', 'left', 'right', 'outer').

    Ejemplo:
        unir_hojas_excel(
            "datos.xlsx", 
            "Clientes", "Pedidos", 
            "ID_Cliente", "Cliente_ID", 
            "Clientes_Pedidos", 
            tipo_union="left"
        )
    """
    try:
        # === Cargar las dos hojas ===
        df1 = pd.read_excel(ruta_archivo, sheet_name=hoja_1, engine="openpyxl")
        df2 = pd.read_excel(ruta_archivo, sheet_name=hoja_2, engine="openpyxl")

        # === Verificar que las columnas existan ===
        if campo_hoja_1 not in df1.columns:
            raise ValueError(f"La columna '{campo_hoja_1}' no existe en la hoja '{hoja_1}'.")

        if campo_hoja_2 not in df2.columns:
            raise ValueError(f"La columna '{campo_hoja_2}' no existe en la hoja '{hoja_2}'.")

        # === Unir las hojas ===
        df_unido = pd.merge(df1, df2, how=tipo_union, left_on=campo_hoja_1, right_on=campo_hoja_2)

        # === Cargar el libro existente ===
        book = load_workbook(ruta_archivo)

        # Si la hoja destino ya existe, eliminarla
        if hoja_destino in book.sheetnames:
            del book[hoja_destino]

        # === Guardar la nueva hoja sin alterar las demás ===
        with pd.ExcelWriter(ruta_archivo, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            writer._book = book
            df_unido.to_excel(writer, sheet_name=hoja_destino, index=False)

        print(f"✅ Se unieron las hojas '{hoja_1}' y '{hoja_2}' en la nueva hoja '{hoja_destino}' ({len(df_unido)} registros). \nTipo: {tipo_union}")
    except Exception as error:
            print(f"Error en método unir_hojas_excel: {error}")

def unir_hojas_excel_cadena(ruta_archivo: str, operaciones_union: list):
    """
    Optimiza recursos abriendo el archivo Excel UNA SOLA VEZ.
    Realiza múltiples uniones encadenadas en memoria RAM y guarda todo al final.

    Parámetros:
        ruta_archivo (str): Ruta del archivo Excel (.xlsx).
        operaciones_union (list): Lista de diccionarios con la configuración de cada unión.
    """
    if not os.path.exists(ruta_archivo) or os.path.getsize(ruta_archivo) == 0:
        print(f"❌ Error: El archivo '{ruta_archivo}' no existe o está vacío.")
        return

    try:
        print(f"📂 Cargando archivo en memoria para uniones en lote: {ruta_archivo}")
        
        # 1. Leer TODAS las hojas del archivo de una sola vez para tenerlas en memoria
        # Esto devuelve un diccionario donde la clave es el nombre de la hoja y el valor es su DataFrame
        diccionario_hojas = pd.read_excel(ruta_archivo, sheet_name=None, engine="openpyxl")
        
        # 2. Procesar cada unión secuencialmente en la memoria RAM
        for paso, op in enumerate(operaciones_union, start=1):
            h1 = op['hoja_1']
            h2 = op['hoja_2']
            c1 = op['campo_hoja_1']
            c2 = op['campo_hoja_2']
            destino = op['hoja_destino']
            tipo = op.get('tipo_union', 'inner')

            print(f"🔄 Paso {paso}: Uniendo '{h1}' + '{h2}' -> '{destino}' ({tipo})...")

            # Verificar que las hojas existan en nuestro mapa de memoria
            if h1 not in diccionario_hojas:
                print(f"  ❌ Error: La hoja '{h1}' no existe en memoria. ¿Olvidaste el orden de la cadena?")
                return
            if h2 not in diccionario_hojas:
                print(f"  ❌ Error: La hoja '{h2}' no existe en el archivo.")
                return

            df1 = diccionario_hojas[h1]
            df2 = diccionario_hojas[h2]

            # Verificar columnas
            if c1 not in df1.columns:
                print(f"  ❌ Error: La columna '{c1}' no existe en la hoja '{h1}'. Columns: {list(df1.columns)}")
                return
            if c2 not in df2.columns:
                print(f"  ❌ Error: La columna '{c2}' no existe en la hoja '{h2}'. Columns: {list(df2.columns)}")
                return

            # Realizar la unión en la RAM
            df_resultado = pd.merge(df1, df2, how=tipo, left_on=c1, right_on=c2)
            
            # ⚠️ CLAVE DE LA OPTIMIZACIÓN: Guardamos el resultado inmediatamente en nuestro mapa de memoria.
            # Esto permite que la siguiente unión del bucle pueda usar 'destino' como origen sin haber ido al disco.
            diccionario_hojas[destino] = df_resultado
            print(f"  ✅ Unión completada en memoria. {len(df_resultado)} registros generados.")

        # 3. Guardar absolutamente todas las hojas de regreso al disco en un solo flujo seguro
        print(f"💾 Escribiendo todas las hojas actualizadas en el archivo...")
        with pd.ExcelWriter(ruta_archivo, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            # Iteramos solo sobre las nuevas hojas generadas por las operaciones de unión
            for op in operaciones_union:
                destino = op['hoja_destino']
                diccionario_hojas[destino].to_excel(writer, sheet_name=destino, index=False)
                
        print(f"🎉 ¡Todas las uniones en cadena se guardaron correctamente en '{ruta_archivo}'!")

    except zipfile.BadZipFile:
        print(f"❌ Error crítico: El archivo '{ruta_archivo}' está corrupto o incompleto (BadZipFile).")
    except PermissionError:
        print(f"❌ Error de permisos: El archivo '{ruta_archivo}' está abierto en Excel. Ciérralo.")
    except Exception as error:
        print(f"❌ Ocurrió un error inesperado en la cadena de uniones: {error}")
