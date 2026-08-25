import os
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def validar_formato_nombre_apellido(
    ruta_archivo: str,
    nombre_hoja: str,
    columna_a_validar: str,
    nombre_columna_resultado: str = "Validación Formato",
    ruta_salida: str = None
):
    """
    Valida que las celdas de una columna específica cumplan el formato 'nombre.apellido'
    en una hoja de un archivo Excel existente. Agrega una nueva columna con el resultado.
    
    Parámetros:
        ruta_archivo (str): Ruta del archivo Excel existente.
        nombre_hoja (str): Nombre de la hoja donde está la columna a validar.
        columna_a_validar (str): Nombre de la columna que se desea validar.
        nombre_columna_resultado (str): Nombre de la nueva columna a agregar.
        ruta_salida (str): Ruta del archivo de salida. Si no se especifica, se sobreescribe el original.
    
    Retorna:
        None
    """

    # === Expresión regular para el formato nombre.apellido ===
    patron = re.compile(r'^[a-zA-Z]+\.{1}[a-zA-Z]+$')

    # === Cargar solo la columna a validar ===
    df_col = pd.read_excel(
        ruta_archivo,
        sheet_name=nombre_hoja,
        usecols=[columna_a_validar],
        engine="openpyxl"
    )

    # === Validar formato ===
    df_col[nombre_columna_resultado] = df_col[columna_a_validar].apply(
        lambda valor: "Cumple" if isinstance(valor, str) and re.match(patron, valor.strip()) else "No cumple"
    )

    # === Cargar libro completo con openpyxl ===
    wb = load_workbook(ruta_archivo)
    ws = wb[nombre_hoja]

    # === Agregar la nueva columna al final ===
    # Encontrar la siguiente columna vacía
    col_final = ws.max_column + 1
    ws.cell(row=1, column=col_final, value=nombre_columna_resultado)

    # Escribir resultados fila por fila (comenzando en la fila 2)
    for i, valor in enumerate(df_col[nombre_columna_resultado], start=2):
        ws.cell(row=i, column=col_final, value=valor)

    # === Guardar cambios ===
    if ruta_salida is None:
        ruta_salida = ruta_archivo  # Sobrescribir
    wb.save(ruta_salida)

    print(f"✅ Validación completada. Archivo guardado como: {ruta_salida}")

def validar_formato_nombre_apellido(
    ruta_archivo: str,
    nombre_hoja: str,
    columna_objetivo: str,
    expresion_regular: str = r"^[A-Za-z]+\.[A-Za-z]+$"
):
    """
    Valida una columna en una hoja de Excel para verificar que cumpla el formato:
    'primerNombre.segundoNombre' (solo letras sin tilde, sin ñ).
    Las celdas que no cumplan se marcan con color rojo claro.

    Parámetros:
        ruta_archivo (str): Ruta del archivo Excel.
        nombre_hoja (str): Nombre de la hoja a validar.
        columna_objetivo (str): Letra de la columna (por ejemplo, 'B').
    """

    # Expresión regular: primerNombre.segundoNombre
    # Solo letras (A-Z o a-z), sin tildes ni ñ
    patron = re.compile(expresion_regular)

    # Cargar el archivo
    wb = load_workbook(ruta_archivo)
    if nombre_hoja not in wb.sheetnames:
        raise ValueError(f"La hoja '{nombre_hoja}' no existe en el archivo.")

    ws = wb[nombre_hoja]

    # Color de fondo (rojo claro)
    rojo_claro = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Contador de errores
    errores = 0

    # Iterar por las filas de la columna (saltando encabezado si hay)
    for fila in range(2, ws.max_row + 1):
        celda = ws[f"{columna_objetivo}{fila}"]
        valor = str(celda.value).strip() if celda.value else ""

        if not patron.match(valor):
            celda.fill = rojo_claro
            errores += 1

    # Guardar los cambios
    wb.save(ruta_archivo)
    print(f"✅ Validación completada. {errores} celdas no cumplen el formato y fueron marcadas en rojo.")

def validar_y_formato_celda_con_color(
    archivo_ruta: str, 
    hoja_nombre: str, 
    columna_identificador: str,
    patron_condicion: str = r'^[a-z]+\.[a-z]+$'
):
    """
    Valida el formato 'nombre.apellido' en una hoja, colorea las celdas no válidas
    y sobrescribe SOLO esa hoja en el archivo, manteniendo el resto de hojas intactas.

    Args:
        archivo_ruta (str): Ruta y nombre del archivo Excel a procesar.
        hoja_nombre (str): Nombre de la hoja a analizar y sobrescribir.
        columna_identificador (str): Nombre de la columna que contiene los identificadores.
    """
    
    # Patrón de regex para "nombre.apellido" (solo minúsculas)
    REGEX_IDENTIFICADOR = r'^[a-z]+\.[a-z]+$'
    
    # Color Rojo Claro para el fondo de las celdas no válidas
    COLOR_NO_VALIDO = 'background-color: #FFCDD2' 
    COLOR_DEFAULT = ''

    try:
        # 1. Cargar TODAS las hojas del archivo
        print(f"⚙️  Cargando todas las hojas del archivo '{archivo_ruta}'...")
        xls = pd.ExcelFile(archivo_ruta)
        # Diccionario con todas las hojas: {'NombreHoja': DataFrame}
        diccionario_hojas = pd.read_excel(xls, sheet_name=xls.sheet_names)
        
        if hoja_nombre not in diccionario_hojas:
            print(f"❌ Error: La hoja '{hoja_nombre}' no se encontró en el archivo.")
            return

        df_a_validar = diccionario_hojas[hoja_nombre]
        print(f"✅ Hoja '{hoja_nombre}' cargada. Total de filas: {len(df_a_validar)}")
        
    except FileNotFoundError:
        print(f"❌ Error: Archivo no encontrado en la ruta: {archivo_ruta}")
        return
    except KeyError as e:
        print(f"❌ Error: Problema al cargar una hoja o columna: {e}")
        return
    except Exception as e:
        print(f"❌ Error al cargar el archivo/hoja: {e}")
        return

    # 2. Definir la función de estilizado para cada celda
    def resaltar_celdas_no_validas(columna):
        """Devuelve un estilo de color para cada celda de la columna, basado en la validación."""
        estilos = []
        try:
            for valor in columna:
                identificador_valor = str(valor).strip()
                # Validación: fullmatch asegura que toda la cadena cumpla el patrón
                if re.fullmatch(patron_condicion, identificador_valor):
                    estilos.append(COLOR_DEFAULT)
                else:
                    estilos.append(COLOR_NO_VALIDO)
            return estilos
        except KeyError:
            # Si la columna no existe en el subconjunto, no aplicar nada
            return [''] * len(columna)

    # 3. Aplicar el estilo al DataFrame de la hoja de destino
    try:
        # El resultado es un objeto Styler que guarda la información del color
        df_estilizado = df_a_validar.style.apply(
            resaltar_celdas_no_validas, 
            subset=[columna_identificador], 
            axis=0 
        )

    except KeyError:
        print(f"❌ Error: La columna '{columna_identificador}' no se encontró en la hoja '{hoja_nombre}'.")
        return
    
    # 4. Guardar TODAS las hojas de nuevo (Sobreescritura con preservación)
    print(f"⚙️  Sobrescribiendo la hoja '{hoja_nombre}' y preservando el resto...")
    
    try:
        # Usamos mode='w' (write/escribir) con engine='xlsxwriter' (necesario para estilos).
        # Esto sobrescribe el archivo, por lo que debemos reescribir todas las hojas.
        with pd.ExcelWriter(archivo_ruta, engine='xlsxwriter', mode='w') as writer:
            
            # Recorrer todas las hojas originales
            for nombre_hoja, df in diccionario_hojas.items():
                if nombre_hoja == hoja_nombre:
                    # Si es la hoja a modificar, escribimos el objeto Styler
                    df_estilizado.to_excel(writer, sheet_name=nombre_hoja, index=False)
                else:
                    # Si es cualquier otra hoja, la reescribimos sin modificar
                    df.to_excel(writer, sheet_name=nombre_hoja, index=False)
        
        print("\n" + "-" * 60)
        print("🎉 Éxito en la validación y sobrescritura:")
        print(f"   La hoja '{hoja_nombre}' ha sido validada, estilizada y sobrescrita.")
        print(f"   El resto de hojas se mantienen intactas.")
        print(f"   Archivo modificado: {os.path.abspath(archivo_ruta)}")
        print("-" * 60)

    except Exception as e:
        print(f"❌ Error al escribir en el archivo: {e}")


# ## --- Ejemplo de Uso de la Función ---

# if __name__ == '__main__':
#     # ⚠️ Asegúrate de que el archivo exista y ten a mano los nombres exactos.
    
#     ARCHIVO_MAESTRO = 'Usuarios_Reporte.xlsx'
#     HOJA_LECTURA = 'Cuentas'
#     COLUMNA_ID = 'Identificador_Cuenta'
    
#     print("Iniciando la validación con sobrescritura...")
    
#     # Requisitos: pip install pandas openpyxl xlsxwriter jinja2
    
#     validar_y_formato_celda_con_color(
#         archivo_ruta=ARCHIVO_MAESTRO,
#         hoja_nombre=HOJA_LECTURA,
#         columna_identificador=COLUMNA_ID
#     )