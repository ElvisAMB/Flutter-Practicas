import os
import pandas as pd
from openpyxl import load_workbook


def ordenar_hoja_excel(
    ruta_archivo: str,
    hoja_origen: str,
    campo_orden: str,
    hoja_destino: str = None,
    ascendente: bool = True,
):
    """
    Ordena los datos de una hoja específica de un archivo Excel
    según un campo determinado y guarda el resultado en la misma hoja
    o en una nueva hoja, sin alterar las demás.

    Parámetros:
        ruta_archivo (str): Ruta del archivo Excel existente.
        hoja_origen (str): Nombre de la hoja a ordenar.
        campo_orden (str): Nombre del campo (columna) por el cual ordenar.
        hoja_destino (str): (Opcional) Nombre de la hoja donde guardar el resultado.
                            Si no se especifica, se sobreescribe la hoja original.
        ascendente (bool): True para orden ascendente, False para descendente.
    """

    # === Leer la hoja específica ===
    df = pd.read_excel(ruta_archivo, sheet_name=hoja_origen, engine="openpyxl")

    # === Validar que la columna exista ===
    if campo_orden not in df.columns:
        raise ValueError(
            f"La columna '{campo_orden}' no existe en la hoja '{hoja_origen}'."
        )

    # === Ordenar la información ===
    df_ordenado = df.sort_values(by=campo_orden, ascending=ascendente)

    # === Cargar el archivo existente ===
    book = load_workbook(ruta_archivo)

    # Si no se indica hoja destino, se usa la misma hoja
    hoja_destino = hoja_destino or hoja_origen

    # Si la hoja destino existe, eliminarla (para reemplazar)
    if hoja_destino in book.sheetnames:
        del book[hoja_destino]

    # === Escribir el DataFrame ordenado en la hoja destino ===
    with pd.ExcelWriter(
        ruta_archivo, engine="openpyxl", mode="a", if_sheet_exists="replace"
    ) as writer:
        writer._book = book
        df_ordenado.to_excel(writer, sheet_name=hoja_destino, index=False)

    print(
        f"✅ La hoja '{hoja_origen}' fue ordenada por '{campo_orden}' y guardada como '{hoja_destino}'."
    )


def ordenar_contenido_hoja_excel(
    archivo_ruta: str,
    hoja_a_ordenar: str,
    campos_ordenacion: list,
    ascendente: bool = True,
):
    """
    Carga todas las hojas de un archivo Excel, ordena una hoja específica
    por múltiples campos y sobrescribe esa hoja en el archivo, manteniendo
    el contenido de las demás hojas intacto.

    Args:
        archivo_ruta (str): Ruta del archivo Excel.
        hoja_a_ordenar (str): Nombre de la hoja que se desea ordenar.
        campos_ordenacion (list): Lista de strings con los nombres exactos de las columnas para ordenar.
        ascendente (bool): Si es True, ordena ascendentemente; si es False, descendentemente.
    """

    try:
        print(f"⚙️  Cargando todas las hojas del archivo '{archivo_ruta}'...")
        # 1. Cargar TODAS las hojas del archivo en un diccionario de DataFrames
        xls = pd.ExcelFile(archivo_ruta)
        diccionario_hojas = pd.read_excel(xls, sheet_name=xls.sheet_names)

        if hoja_a_ordenar not in diccionario_hojas:
            print(f"❌ Error: La hoja '{hoja_a_ordenar}' no se encontró en el archivo.")
            return

        # Obtener el DataFrame de la hoja a modificar
        df_a_modificar = diccionario_hojas[hoja_a_ordenar]

        # 2. Ordenar el DataFrame
        print(
            f"⚙️  Ordenando la hoja '{hoja_a_ordenar}' por los campos: {campos_ordenacion}..."
        )

        # El método .sort_values acepta una lista de columnas y un valor booleano o lista de booleanos.
        df_ordenado = df_a_modificar.sort_values(
            by=campos_ordenacion,
            ascending=ascendente,
            ignore_index=True,  # Opcional: restablece el índice después de ordenar
        )

        # Reemplazar el DataFrame original en el diccionario con el DataFrame ordenado
        diccionario_hojas[hoja_a_ordenar] = df_ordenado

        # 3. Guardar TODAS las hojas de nuevo en el archivo
        # Esto sobrescribe el archivo, pero como se han cargado todas las hojas,
        # el contenido de las hojas no modificadas se mantiene.
        print(
            f"⚙️  Sobrescribiendo el archivo con la hoja '{hoja_a_ordenar}' ordenada..."
        )

        with pd.ExcelWriter(archivo_ruta, engine="openpyxl", mode="w") as writer:
            for nombre_hoja, df in diccionario_hojas.items():
                df.to_excel(writer, sheet_name=nombre_hoja, index=False)

        print("\n" + "-" * 60)
        print("🎉 Éxito en la ordenación:")
        print(f"   La hoja '{hoja_a_ordenar}' ha sido ordenada y sobrescrita.")
        print(f"   Archivo modificado: {os.path.abspath(archivo_ruta)}")
        print("-" * 60)

    except FileNotFoundError:
        print(f"❌ Error: Archivo no encontrado en la ruta: {archivo_ruta}")
    except KeyError as e:
        print(f"❌ Error: Uno o más campos de ordenación no existen en la hoja: {e}")
    except Exception as e:
        print(f"❌ Ocurrió un error inesperado: {e}")

def ordenar_hojas_excel(
    archivo_ruta: str,
    configuraciones_ordenamiento: list[dict],
):
    """
    Ordena múltiples hojas de un archivo Excel en una sola operación,
    manteniendo las demás hojas sin modificaciones.

    Parámetros:
        archivo_ruta (str):
            Ruta completa del archivo Excel.

        configuraciones_ordenamiento (list[dict]):
            Lista de configuraciones. Cada elemento debe contener:

            {
                "hoja": "NombreHoja",
                "campos": ["Campo1", "Campo2"],
                "ascendente": True
            }

            También puede utilizarse una lista de booleanos en
            "ascendente" para definir el sentido de cada campo:

            {
                "hoja": "NombreHoja",
                "campos": ["Sucursal", "Nombre"],
                "ascendente": [True, False]
            }

    Ejemplo:
        ordenar_hojas_excel(
            archivo_ruta="archivo.xlsx",
            configuraciones_ordenamiento=[
                {
                    "hoja": "UsrNomina",
                    "campos": ["Sucursal", "Nombre"],
                    "ascendente": True
                },
                {
                    "hoja": "AccountsEnabled",
                    "campos": ["Tipo", "Sucursal", "Primer apellido"],
                    "ascendente": True
                }
            ]
        )
    """

    if not os.path.isfile(archivo_ruta):
        raise FileNotFoundError(f"No existe el archivo: {archivo_ruta}")

    if not configuraciones_ordenamiento:
        raise ValueError(
            "Debe proporcionar al menos una configuración de ordenamiento."
        )

    workbook = None

    try:
        print(f"⚙️ Abriendo archivo: {archivo_ruta}")

        # Una sola apertura del archivo
        workbook = load_workbook(filename=archivo_ruta)

        # Procesar todas las hojas solicitadas
        for configuracion in configuraciones_ordenamiento:

            nombre_hoja = configuracion["hoja"]
            campos = configuracion["campos"]
            ascendente = configuracion.get("ascendente", True)

            print(f"⚙️ Procesando hoja '{nombre_hoja}'...")

            # ---------------------------------------------------------
            # Validar hoja
            # ---------------------------------------------------------
            if nombre_hoja not in workbook.sheetnames:
                raise KeyError(f"La hoja '{nombre_hoja}' no existe en el archivo.")

            if not campos:
                raise ValueError(
                    f"No se definieron campos para ordenar " f"la hoja '{nombre_hoja}'."
                )

            hoja = workbook[nombre_hoja]

            # ---------------------------------------------------------
            # Obtener encabezados
            # ---------------------------------------------------------
            encabezados = {
                celda.value: indice
                for indice, celda in enumerate(hoja[1], start=1)
                if celda.value is not None
            }

            # ---------------------------------------------------------
            # Validar campos
            # ---------------------------------------------------------
            campos_no_existentes = [
                campo for campo in campos if campo not in encabezados
            ]

            if campos_no_existentes:
                raise KeyError(
                    f"En la hoja '{nombre_hoja}' no existen "
                    f"los siguientes campos: "
                    f"{campos_no_existentes}"
                )

            # ---------------------------------------------------------
            # Obtener todas las filas de datos
            # ---------------------------------------------------------
            filas = list(hoja.iter_rows(min_row=2, values_only=False))

            if not filas:
                print(
                    f"   ℹ️ La hoja '{nombre_hoja}' no contiene " f"filas para ordenar."
                )
                continue

            # ---------------------------------------------------------
            # Normalizar ascendente
            # ---------------------------------------------------------
            if isinstance(ascendente, bool):
                ordenes = [ascendente] * len(campos)

            elif isinstance(ascendente, list):
                if len(ascendente) != len(campos):
                    raise ValueError(
                        f"La cantidad de valores de 'ascendente' "
                        f"debe coincidir con la cantidad de campos "
                        f"en la hoja '{nombre_hoja}'."
                    )

                ordenes = ascendente

            else:
                raise TypeError("'ascendente' debe ser bool o list[bool].")

            # ---------------------------------------------------------
            # Ordenamiento estable de múltiples campos
            # ---------------------------------------------------------
            #
            # Se ordena desde el último campo hacia el primero.
            # Esto permite utilizar un sentido diferente por columna.
            #
            for campo, es_ascendente in reversed(list(zip(campos, ordenes))):

                indice_columna = encabezados[campo]

                filas.sort(
                    key=lambda fila: (
                        fila[indice_columna - 1].value
                        if fila[indice_columna - 1].value is not None
                        else ""
                    ),
                    reverse=not es_ascendente,
                )

            # ---------------------------------------------------------
            # Reescribir únicamente las filas de la hoja
            # ---------------------------------------------------------
            for numero_fila, fila in enumerate(filas, start=2):
                for numero_columna, celda_origen in enumerate(fila, start=1):
                    hoja.cell(row=numero_fila, column=numero_columna).value = (
                        celda_origen.value
                    )

            print(f"   ✅ Hoja '{nombre_hoja}' ordenada correctamente.")

        # -------------------------------------------------------------
        # Guardar UNA SOLA VEZ
        # -------------------------------------------------------------
        print("💾 Guardando cambios...")

        workbook.save(archivo_ruta)

        print("\n" + "-" * 60)
        print("🎉 Proceso completado correctamente.")
        print(f"📄 Archivo modificado: " f"{os.path.abspath(archivo_ruta)}")
        print("-" * 60)

    except Exception as e:

        print(f"❌ Error durante el procesamiento del archivo: {e}")

        raise

    finally:

        # Liberar correctamente el recurso
        if workbook is not None:
            workbook.close()

            print("🔒 Archivo cerrado correctamente.")
